from __future__ import annotations

import argparse
import math
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font


SIGNIFICANT_COLS = [
    "B",
    "D",
    "H",
    "L",
    "M",
    "N",
    "O",
    "Q",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AJ",
    "AK",
    "AL",
    "AM",
    "AN",
    "AO",
]

PAGE_LAYOUTS = {
    "general": {
        "template_name": "__TEMPLATE_PAGE_GENERAL",
        "index_link_cell": "Q1",
        "page_count_cell": "P36",
        "timestamp_cell": "A37",
        "data_start": 4,
        "data_end": 35,
    },
    "ttc": {
        "template_name": "__TEMPLATE_PAGE_TTC",
        "index_link_cell": "S1",
        "page_count_cell": "Q21",
        "timestamp_cell": "A22",
        "data_start": 4,
        "data_end": 20,
    },
    "measurement": {
        "template_name": "__TEMPLATE_PAGE_MEASURE",
        "index_link_cell": "V1",
        "page_count_cell": "U32",
        "timestamp_cell": "A33",
        "data_start": 4,
        "data_end": 31,
    },
}


@dataclass
class Record:
    source_title: str
    source_file: str
    layout: str
    pay_item: str
    description: str
    unit: str | None
    number_required: str | int | float | None
    duration_days: str | int | float | None
    frequency_days: str | int | float | None
    length_ft: str | int | float | None
    width_ft: str | int | float | None
    qty_p: str | int | float | None
    qty_f: str | int | float | None
    total_p: str | int | float | None
    total_f: str | int | float | None
    secondary_unit: str | None
    secondary_p: str | int | float | None
    secondary_f: str | int | float | None
    const_phase: str | None
    wall_no: str | int | float | None
    site_no: str | int | float | None
    alignment: str | None
    begin_station: str | None
    end_station: str | None
    location_description: str | None
    offset: str | int | float | None
    side: str | None
    element_id: str | int | float | None
    design_notes: str | None
    construction_remarks: str | None
    show_context: bool = False
    page_no: int | None = None


@dataclass
class ItemGroup:
    pay_item: str
    description: str
    records: list[Record] = field(default_factory=list)


@dataclass
class SourceSummary:
    title: str
    filename: str
    layout: str
    groups: list[ItemGroup]


def normalize_space(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def cleaned_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = normalize_space(value)
        return text or None
    return value


def numeric_sort_key(pay_item: str):
    parts = re.findall(r"\d+|[A-Za-z]+", pay_item)
    key = []
    for part in parts:
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part.lower()))
    return key


def safe_sheet_title(base: str, seen: set[str]) -> str:
    title = re.sub(r"[:\\/?*\[\]]", " ", base)
    title = re.sub(r"\s+", " ", title).strip()
    if not title:
        title = "Sheet"
    title = title[:31]
    candidate = title
    counter = 2
    while candidate in seen:
        suffix = f" {counter}"
        candidate = f"{title[:31-len(suffix)]}{suffix}"
        counter += 1
    seen.add(candidate)
    return candidate


def set_internal_link(cell: Cell, target: str):
    cell.hyperlink = f"#{target}"
    font = copy(cell.font) if cell.font else Font()
    font.underline = "single"
    font.color = "0000FF"
    cell.font = font


def clear_range(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if cell.coordinate in ws.merged_cells:
                continue
            cell.value = None
            cell.hyperlink = None


def clone_row_style(ws, source_row: int, target_row: int, start_col: int, end_col: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(start_col, end_col + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)
        target._style = copy(source._style)
        if source.number_format:
            target.number_format = copy(source.number_format)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def ensure_detail_rows(ws, needed_last_row: int):
    if needed_last_row <= ws.max_row:
        return
    source_row = 6
    start = ws.max_row + 1
    amount = needed_last_row - ws.max_row
    ws.insert_rows(start, amount)
    for row_idx in range(start, needed_last_row + 1):
        clone_row_style(ws, source_row, row_idx, 1, 13)


def detect_layout(title: str, groups: list[ItemGroup]) -> str:
    title_lower = title.lower()
    if "temporary traffic control" in title_lower:
        return "ttc"
    for group in groups:
        for record in group.records:
            if any(record.__dict__[field] for field in ("length_ft", "width_ft")):
                return "measurement"
    return "general"


def parse_source_file(path: Path) -> SourceSummary:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    title = normalize_space(ws["A1"].value or ws["B1"].value or path.stem)

    groups: list[ItemGroup] = []
    current_group: ItemGroup | None = None
    context = {
        "pay_item": "",
        "description": "",
        "unit": None,
        "number_required": None,
        "duration_days": None,
        "frequency_days": None,
        "length_ft": None,
        "width_ft": None,
        "total_p": None,
        "total_f": None,
        "secondary_unit": None,
        "secondary_p": None,
        "secondary_f": None,
        "design_notes": None,
        "construction_remarks": None,
    }

    for row_idx in range(4, ws.max_row + 1):
        if not any(cleaned_value(ws[f"{col}{row_idx}"].value) is not None for col in SIGNIFICANT_COLS):
            continue

        pay_item_value = cleaned_value(ws[f"B{row_idx}"].value)
        description_value = cleaned_value(ws[f"D{row_idx}"].value)
        if pay_item_value is not None:
            context["pay_item"] = normalize_space(pay_item_value)
        if description_value is not None:
            context["description"] = normalize_space(description_value)

        for field, col in {
            "unit": "H",
            "number_required": "L",
            "duration_days": "M",
            "frequency_days": "N",
            "length_ft": "O",
            "width_ft": "Q",
            "total_p": "Y",
            "total_f": "Z",
            "secondary_unit": "AA",
            "secondary_p": "AB",
            "secondary_f": "AC",
            "design_notes": "AN",
            "construction_remarks": "AO",
        }.items():
            value = cleaned_value(ws[f"{col}{row_idx}"].value)
            if value is not None:
                context[field] = value

        if not context["pay_item"]:
            continue

        if pay_item_value is not None:
            current_group = ItemGroup(
                pay_item=context["pay_item"],
                description=context["description"] or context["pay_item"],
            )
            groups.append(current_group)

        if current_group is None:
            current_group = ItemGroup(
                pay_item=context["pay_item"],
                description=context["description"] or context["pay_item"],
            )
            groups.append(current_group)

        record = Record(
            source_title=title,
            source_file=path.name,
            layout="",
            pay_item=context["pay_item"],
            description=context["description"],
            unit=context["unit"],
            number_required=cleaned_value(ws[f"L{row_idx}"].value),
            duration_days=cleaned_value(ws[f"M{row_idx}"].value),
            frequency_days=cleaned_value(ws[f"N{row_idx}"].value),
            length_ft=cleaned_value(ws[f"O{row_idx}"].value),
            width_ft=cleaned_value(ws[f"Q{row_idx}"].value),
            qty_p=cleaned_value(ws[f"W{row_idx}"].value),
            qty_f=cleaned_value(ws[f"X{row_idx}"].value),
            total_p=cleaned_value(ws[f"Y{row_idx}"].value),
            total_f=cleaned_value(ws[f"Z{row_idx}"].value),
            secondary_unit=context["secondary_unit"],
            secondary_p=cleaned_value(ws[f"AB{row_idx}"].value),
            secondary_f=cleaned_value(ws[f"AC{row_idx}"].value),
            const_phase=cleaned_value(ws[f"AD{row_idx}"].value),
            wall_no=cleaned_value(ws[f"AE{row_idx}"].value),
            site_no=cleaned_value(ws[f"AF{row_idx}"].value),
            alignment=cleaned_value(ws[f"AG{row_idx}"].value),
            begin_station=cleaned_value(ws[f"AH{row_idx}"].value),
            end_station=cleaned_value(ws[f"AI{row_idx}"].value),
            location_description=cleaned_value(ws[f"AJ{row_idx}"].value),
            offset=cleaned_value(ws[f"AK{row_idx}"].value),
            side=cleaned_value(ws[f"AL{row_idx}"].value),
            element_id=cleaned_value(ws[f"AM{row_idx}"].value),
            design_notes=cleaned_value(ws[f"AN{row_idx}"].value) or context["design_notes"],
            construction_remarks=cleaned_value(ws[f"AO{row_idx}"].value) or context["construction_remarks"],
            show_context=len(current_group.records) == 0,
        )
        current_group.records.append(record)

    layout = detect_layout(title, groups)
    for group in groups:
        for record in group.records:
            record.layout = layout
    return SourceSummary(title=title, filename=path.name, layout=layout, groups=groups)


def paginate_summary(summary: SourceSummary) -> list[list[Record]]:
    config = PAGE_LAYOUTS[summary.layout]
    capacity = config["data_end"] - config["data_start"] + 1
    pages: list[list[Record]] = []
    current_page: list[Record] = []

    for group in summary.groups:
        group_records = [copy_record(record) for record in group.records]
        if len(group_records) > capacity:
            if current_page:
                pages.append(current_page)
                current_page = []
            start = 0
            while start < len(group_records):
                chunk = [copy_record(record) for record in group_records[start : start + capacity]]
                if start > 0 and chunk:
                    chunk[0].show_context = True
                pages.append(chunk)
                start += capacity
            continue

        if current_page and len(current_page) + len(group_records) > capacity:
            pages.append(current_page)
            current_page = []

        current_page.extend(group_records)

    if current_page:
        pages.append(current_page)

    return pages


def copy_record(record: Record) -> Record:
    return Record(**record.__dict__)


def page_row_values(record: Record):
    if record.layout == "ttc":
        return {
            "A": record.pay_item if record.show_context else None,
            "B": record.description if record.show_context else None,
            "D": record.unit if record.show_context else None,
            "E": record.number_required,
            "F": record.duration_days,
            "G": record.qty_p,
            "H": record.qty_f,
            "I": record.total_p if record.show_context else None,
            "J": record.total_f if record.show_context else None,
            "K": record.secondary_unit if record.show_context else None,
            "L": record.secondary_p,
            "M": record.secondary_f,
            "N": record.const_phase or record.alignment,
            "O": record.design_notes,
            "P": record.construction_remarks,
        }

    if record.layout == "measurement":
        return {
            "A": record.pay_item if record.show_context else None,
            "B": record.description if record.show_context else None,
            "D": record.unit if record.show_context else None,
            "E": record.length_ft,
            "F": record.width_ft,
            "G": record.qty_p,
            "H": record.qty_f,
            "I": record.total_p if record.show_context else None,
            "J": record.total_f if record.show_context else None,
            "K": record.secondary_unit if record.show_context else None,
            "L": record.secondary_p,
            "M": record.secondary_f,
            "N": record.alignment or record.const_phase,
            "O": record.begin_station,
            "P": record.end_station,
            "Q": record.side,
            "R": record.element_id,
            "S": record.design_notes,
            "T": record.construction_remarks,
        }

    return {
        "A": record.pay_item if record.show_context else None,
        "B": record.description if record.show_context else None,
        "D": record.unit if record.show_context else None,
        "E": record.qty_p,
        "F": record.qty_f,
        "G": record.total_p if record.show_context else None,
        "H": record.total_f if record.show_context else None,
        "I": record.alignment or record.const_phase,
        "J": record.begin_station or record.number_required,
        "K": record.end_station or record.duration_days,
        "L": record.side or record.frequency_days,
        "M": record.element_id or record.location_description,
        "N": record.design_notes
        or (
            f"Secondary: {record.secondary_unit or ''} {record.secondary_p or ''} {record.secondary_f or ''}".strip()
            if any(v not in (None, "") for v in (record.secondary_unit, record.secondary_p, record.secondary_f))
            else None
        ),
        "O": record.construction_remarks,
    }


def write_page_sheet(ws, summary_title: str, page_no: int, total_pages: int, records: list[Record]):
    layout = records[0].layout if records else "general"
    config = PAGE_LAYOUTS[layout]
    data_start = config["data_start"]
    data_end = config["data_end"]

    clear_range(ws, 1, ws.max_row, 1, ws.max_column)
    ws["A1"] = summary_title
    link_cell = ws[config["index_link_cell"]]
    link_cell.value = "Ir a Indice"
    set_internal_link(link_cell, "Index!A1")
    ws[config["page_count_cell"]] = f"Page {page_no} of {total_pages}"
    ws[config["timestamp_cell"]] = datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")

    # Restore static headers that were cleared.
    template_headers = {
        "general": {
            "A2": "Pay Item\nNumber",
            "B2": "Pay Item Description",
            "D2": "Units of\nMeasure",
            "E2": "Quantity",
            "G2": "Total Quantity",
            "I2": "Location",
            "N2": "Design Notes",
            "O2": "Construction Remarks",
            "E3": "P",
            "F3": "F",
            "G3": "P",
            "H3": "F",
            "I3": "Alignment",
            "J3": "Begin\nStation",
            "K3": "End\nStation",
            "L3": "Side",
            "M3": "Element ID",
        },
        "ttc": {
            "A2": "Pay Item\nNumber",
            "B2": "Pay Item Description",
            "D2": "Units of\nMeasure",
            "E2": "Number\nRequired",
            "F2": "Duration\n(Days)",
            "G2": "Quantity",
            "I2": "Total Quantity",
            "K2": "Secondary Quantity",
            "N2": "Location",
            "O2": "Design Notes",
            "P2": "Construction Remarks",
            "G3": "P",
            "H3": "F",
            "I3": "P",
            "J3": "F",
            "K3": "Units",
            "L3": "P",
            "M3": "F",
            "N3": "Const.\nPhase",
        },
        "measurement": {
            "A2": "Pay Item\nNumber",
            "B2": "Pay Item Description",
            "D2": "Units of\nMeasure",
            "E2": "Length (FT)",
            "F2": "Width (FT)",
            "G2": "Quantity",
            "I2": "Total Quantity",
            "K2": "Secondary Quantity",
            "N2": "Location",
            "S2": "Design Notes",
            "T2": "Construction\nRemarks",
            "G3": "P",
            "H3": "F",
            "I3": "P",
            "J3": "F",
            "K3": "Units",
            "L3": "P",
            "M3": "F",
            "N3": "Alignment",
            "O3": "Begin\nStation",
            "P3": "End\nStation",
            "Q3": "Side",
            "R3": "Element\nID",
        },
    }
    for coord, value in template_headers[layout].items():
        ws[coord] = value

    for row_idx in range(data_start, data_end + 1):
        for cell in ws[row_idx]:
            if cell.coordinate in ws.merged_cells:
                continue
            if row_idx in (1, 2, 3):
                continue
            cell.value = None
            cell.hyperlink = None

    for offset, record in enumerate(records):
        target_row = data_start + offset
        for col, value in page_row_values(record).items():
            ws[f"{col}{target_row}"] = value


def prepare_detail_sheet(ws):
    keep_merges = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 4:
            keep_merges.append(str(merged))
        else:
            ws.unmerge_cells(str(merged))

    clear_range(ws, 1, ws.max_row, 1, 14)
    ws.merge_cells("A1:M2")
    for merge in keep_merges:
        if merge != "A1:M2":
            ws.merge_cells(merge)

    ws["A3"] = "Pay Item Number"
    ws["B3"] = "Pay Item Description"
    ws["C3"] = "Units of Measure"
    ws["D3"] = "Quantity"
    ws["F3"] = "Total Quantity"
    ws["H3"] = "Location"

    ws["D4"] = "P"
    ws["E4"] = "F"
    ws["F4"] = "P"
    ws["G4"] = "F"
    ws["H4"] = "Alignment / Phase"
    ws["I4"] = "Begin Sta. / # Req"
    ws["J4"] = "End Sta. / Days"
    ws["K4"] = "Side / Ref"
    ws["L4"] = "Notes"
    ws["M4"] = "Page"

    headers = [
        "pay_item",
        "description",
        "unit",
        "qty_p",
        "qty_f",
        "tot_p",
        "tot_f",
        "loc_1",
        "loc_2",
        "loc_3",
        "ref",
        "notes",
        "page",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(5, col_idx).value = header

    for row_idx in range(6, ws.max_row + 1):
        for col_idx in range(1, 15):
            cell = ws.cell(row_idx, col_idx)
            cell.value = None
            cell.hyperlink = None


def detail_row_values(record: Record):
    loc_1 = record.alignment or record.const_phase or record.location_description
    loc_2 = record.begin_station if record.begin_station is not None else record.number_required
    loc_3 = record.end_station if record.end_station is not None else record.duration_days
    ref = record.side or record.element_id or record.frequency_days or record.wall_no or record.site_no

    notes_parts = []
    for label, value in [
        ("Loc", record.location_description),
        ("Offset", record.offset),
        ("Element", record.element_id),
        ("Wall", record.wall_no),
        ("Site", record.site_no),
        ("Design", record.design_notes),
        ("Remarks", record.construction_remarks),
    ]:
        if value not in (None, ""):
            notes_parts.append(f"{label}: {value}")
    notes = " | ".join(notes_parts)

    return [
        record.pay_item if record.show_context else None,
        record.description if record.show_context else None,
        record.unit if record.show_context else None,
        record.qty_p,
        record.qty_f,
        record.total_p if record.show_context else None,
        record.total_f if record.show_context else None,
        loc_1,
        loc_2,
        loc_3,
        ref,
        notes,
        record.page_no,
    ]


def rebuild_workbook(template_path: Path, source_dir: Path, output_path: Path):
    summaries = [parse_source_file(path) for path in sorted(source_dir.glob("*.xlsx"))]
    summaries.sort(key=lambda s: (normalize_space(s.title).lower(), s.filename.lower()))

    wb = load_workbook(template_path)
    wb["0635 2 11_x0009_PULL & SPLICE BOX, F&"].title = "__TEMPLATE_DETAIL"
    wb["Page 31"].title = "__TEMPLATE_PAGE_GENERAL"
    wb["Page 4"].title = "__TEMPLATE_PAGE_TTC"
    wb["Page 14"].title = "__TEMPLATE_PAGE_MEASURE"

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in {
            "Index",
            "__TEMPLATE_DETAIL",
            "__TEMPLATE_PAGE_GENERAL",
            "__TEMPLATE_PAGE_TTC",
            "__TEMPLATE_PAGE_MEASURE",
        }:
            del wb[sheet_name]

    index_ws = wb["Index"]
    detail_template = wb["__TEMPLATE_DETAIL"]
    page_templates = {
        layout: wb[config["template_name"]] for layout, config in PAGE_LAYOUTS.items()
    }

    page_sheets = []
    all_pages: list[tuple[str, list[Record]]] = []
    for summary in summaries:
        for records in paginate_summary(summary):
            all_pages.append((summary.title, records))

    total_pages = len(all_pages) + 1
    item_to_pages: dict[str, list[int]] = defaultdict(list)
    item_to_desc: dict[str, str] = {}
    item_to_records: dict[str, list[Record]] = defaultdict(list)

    for offset, (summary_title, records) in enumerate(all_pages, start=2):
        layout = records[0].layout if records else "general"
        ws = wb.copy_worksheet(page_templates[layout])
        ws.title = f"Page {offset}"
        write_page_sheet(ws, summary_title, offset, total_pages, records)
        page_sheets.append(ws)

        seen_this_page = set()
        for record in records:
            record.page_no = offset
            item_to_desc.setdefault(record.pay_item, record.description)
            item_to_records[record.pay_item].append(record)
            if record.pay_item not in seen_this_page:
                item_to_pages[record.pay_item].append(offset)
                seen_this_page.add(record.pay_item)

    used_titles = set(wb.sheetnames)
    detail_sheet_names: dict[str, str] = {}
    detail_sheets = []
    for pay_item in sorted(item_to_records, key=numeric_sort_key):
        desc = item_to_desc[pay_item]
        sheet_name = safe_sheet_title(f"{pay_item} {desc}", used_titles)
        detail_sheet_names[pay_item] = sheet_name

        ws = wb.copy_worksheet(detail_template)
        ws.title = sheet_name
        prepare_detail_sheet(ws)
        ws["A1"] = f"Final Measure :{pay_item} {desc}"
        ws["N1"] = "Ir al Indice"
        set_internal_link(ws["N1"], "Index!A1")

        page_links = sorted(set(item_to_pages[pay_item]))
        for link_row in range(2, 2 + max(len(page_links), 1)):
            clone_row_style(ws, 2, link_row, 14, 14)
        for row_idx in range(2, max(15, 2 + len(page_links))):
            ws[f"N{row_idx}"] = None
            ws[f"N{row_idx}"].hyperlink = None
        for idx, page_no in enumerate(page_links, start=2):
            cell = ws[f"N{idx}"]
            cell.value = f"Ir a Pag. {page_no}"
            set_internal_link(cell, f"'Page {page_no}'!A1")

        records = item_to_records[pay_item]
        ensure_detail_rows(ws, max(6, 5 + len(records)))
        for row_idx, record in enumerate(records, start=6):
            for col_idx, value in enumerate(detail_row_values(record), start=1):
                ws.cell(row_idx, col_idx).value = value

        detail_sheets.append(ws)

    # Reset Index and rebuild it.
    clear_range(index_ws, 2, index_ws.max_row, 1, 4)
    template_row_styles = {}
    for col_idx in range(1, 5):
        template_row_styles[col_idx] = copy(index_ws.cell(2, col_idx)._style)
    if len(item_to_records) + 1 > index_ws.max_row:
        start = index_ws.max_row + 1
        amount = len(item_to_records) + 1 - index_ws.max_row
        index_ws.insert_rows(start, amount)
        for row_idx in range(start, start + amount):
            for col_idx in range(1, 5):
                index_ws.cell(row_idx, col_idx)._style = copy(template_row_styles[col_idx])

    for row_idx, pay_item in enumerate(sorted(item_to_records, key=numeric_sort_key), start=2):
        pages = sorted(set(item_to_pages[pay_item]))
        page_text = ", ".join(str(page) for page in pages)
        index_ws[f"A{row_idx}"] = row_idx - 1
        index_ws[f"B{row_idx}"] = pay_item
        index_ws[f"C{row_idx}"] = item_to_desc[pay_item]
        index_ws[f"D{row_idx}"] = page_text
        target = f"'{detail_sheet_names[pay_item]}'!A1"
        set_internal_link(index_ws[f"B{row_idx}"], target)
        set_internal_link(index_ws[f"D{row_idx}"], target)

    # Remove template sheets and set order.
    for template_name in [
        "__TEMPLATE_DETAIL",
        "__TEMPLATE_PAGE_GENERAL",
        "__TEMPLATE_PAGE_TTC",
        "__TEMPLATE_PAGE_MEASURE",
    ]:
        del wb[template_name]

    wb._sheets = [index_ws] + detail_sheets + page_sheets
    wb.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--source-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    rebuild_workbook(args.template, args.source_dir, args.output)


if __name__ == "__main__":
    main()
